import openpyxl
from utils.odpsInit import odpsInit


def load2excel(filepath, begin, end):
    o = odpsInit()
    print(o)
    sql = "select * from stat_table where ds>={begin} and ds<={end}".format(begin=begin, end=end)
    with o.execute_sql(sql).open_reader() as reader:
        print(sql)
        print(reader.count)
        table_title = reader._schema.names
        print(table_title)

        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet('Sheet')
        for i in range(len(table_title)):
            worksheet.cell(1, i + 1, table_title[i])

        write_count = 2
        for r in reader:
            for i in range(len(r.values)):  # 10个依次写入当前行
                worksheet.cell(write_count, i + 1, r.values[i])
            write_count = write_count + 1
        workbook.save(filename=filepath)


if __name__ == '__main__':
    load2excel('xxxx.xlsx', '20180910', '20181009')
