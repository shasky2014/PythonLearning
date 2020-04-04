'''
import xlrd
# xlrd 读xls

import xlwt
# xlwt 写xls

import openpyxl
# 读写 xlsx
'''

import openpyxl

# file = '数据来源.xlsx'
# file = 'd:/PycharmProjects/learn-python/python_coder_share/数据来源.xlsx'
file = '/Users/admin/PycharmProjects/learn-python/python_coder_share/数据来源.xlsx'
wookbook = openpyxl.load_workbook(file, read_only=False, data_only=False)
print(file, wookbook.sheetnames)
excel_sheet = wookbook['Sheet1']
# print(excel_sheet.values)
print([x for x in excel_sheet.values])

sheet_data = [list(x) for x in excel_sheet.values][1:]

print(sheet_data)

for x in sheet_data:
    print(x)

print('-' * 100)


# exit(0)
# 做一个函数，把功能模块化
def read_excel(file, sheet='Sheet1'):
    import openpyxl
    wookbook = openpyxl.load_workbook(file, read_only=False, data_only=False)
    excel_sheet = wookbook[sheet]
    return [list(x) for x in excel_sheet.values][1:]


if __name__ == '__main__':
    file = '/Users/admin/PycharmProjects/learn-python/python_coder_share/数据来源.xlsx'

    excel_data = read_excel(file, sheet='Sheet1')
    print(excel_data)
